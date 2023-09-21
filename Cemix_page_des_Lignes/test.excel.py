import openpyxl
from openpyxl.styles import PatternFill

# Create a new Excel workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Add the column headers to the sheet
column_headers = [
    ['OPERATEUR MIX', 'OPERATEUR ENSACH', 'CLARIST P', 'CLARIST M', 'AIDE MAGASIGNIER'],
    ['date', 'HEURE', 'durée de palette theorique', 'objectif PALETTE theorique', 'TOTAL SAC THEORIQUE'],
    ['TOTAL POID', "total d'heure travail", "durée total d'arret", 'total palette', 'total sac']
]

# Define distinct fill colors for each column group
fill_colors = [
    PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # Yellow
    PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),  # Green
    PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")   # Cyan
]

# Loop through each header row and column, applying formatting
for row_idx, headers in enumerate(column_headers):
    sheet.merge_cells(start_row=row_idx + 1, start_column=1, end_row=row_idx + 1, end_column=len(headers))
    for col_idx, header in enumerate(headers):
        top_left_cell = sheet.cell(row=row_idx + 1, column=col_idx + 1)
        top_left_cell.fill = fill_colors[row_idx]
        
        # Set the value for the top-left cell only (for merged cells)
        if col_idx == 0:
            top_left_cell.value = header

# Save the Excel workbook
workbook.save("formatted_excel.xlsx")




("DECOR TONAL BLANC 105", "bc105", 3),
("DECOR TONAL BLANC 100", "bc100", 3),
("DECOR TONAL BLEU 200", "bl200", 3),
("DECOR TONAL BLEU 205", "bl205", 3),
("DECOR TONAL BRUN 300", "br300", 3),
("DECOR TONAL BRUN 305", "br305", 3),
("DECOR TONAL BRUN 310", "br310", 3),
("DECOR TONAL BRUN 315", "b315", 3),
("DECOR TONAL BRUN 320", "br320", 3),
("DECOR TONAL BRUN 325", "br325", 3),
("DECOR TONAL GRIS 400", "gr400", 3),
("DECOR TONAL GRIS 405", "gr405", 3),
("DECOR TONAL GRIS 410", "gr410", 3),
("DECOR TONAL JAUNE 500", "ja500", 3),
("DECOR TONAL JAUNE 505", "Ja505", 3),
("DECOR TONAL JAUNE 510", "ja510", 3),
("DECOR TONAL JAUNE 515", "tlja515", 3),
("DECOR TONAL JAUNE 520", "ja520", 3),
("DECOR TONAL JAUNE 525", "ja525", 3),
("DECOR TONAL JAUNE 530", "ja530", 3),
("DECOR TONAL JAUNE 535", "ja535", 3),
("DECOR TONAL JAUNE 540", "ja540", 3),
("DECOR TONAL JAUNE 545", "ja545", 3),
("DECOR TONAL JAUNE 550", "ja550", 3),
("DECOR TONAL ORANGE 600", "or600", 3),
("DECOR TONAL ORANGE 605", "or605", 3),
("DECOR TONAL ORANGE 610", "or610", 3),
("DECOR TONAL ORANGE 615", "or615", 3),
("DECOR TONAL ORANGE 620", "or620", 3),
("DECOR TONAL ORANGE 625", "or625", 3),
("DECOR TONAL ROUGE 700", "ro700", 3),
("DECOR TONAL ROUGE 705", "ro705", 3),
("DECOR TONAL ROUGE 710", "ro710", 3),
("DECOR TONAL ROUGE 715", "ro715", 3),
("DECOR TONAL ROUGE 720", "ro720", 3),
("DECOR TONAL ROUGE 725", "ro725", 3),
("DECOR TONAL VERT 800", "vr800", 3),
("DECOR TONAL VERT 805", "vr805", 3),
("DECOR TONAL VERT 810", "tlvr810", 3),
("DECOR TONAL VERT 815",	"vr815", 3),
("DECOR TONAL ECHANTILLON", 	"ech", 3),
("DECOR TONAL BLANC 192",	"bc192", 3),
("DECOR TONAL BRUN 390",	"br390", 3),
("DECOR TONAL GRIS 460",	"gr460", 3),
("DECOR TONAL ORANGE 630",	"ot630", 3),
("DECOR TONAL ROUGE 730",	"ro730", 3),
("DECOR PIETRA GRIS 405",	"gr405", 3),
("DECOR TONAL GRIS 470",	"gr470", 3),
("EXP: DECOR TONAL GRIS 460", 	"gr460", 3),
("DECOR TONAL VERT 860",	"vr860", 3),
("DECOR TONAL GRIS 435",	"gr435", 3),
("DECOR TONAL GRIS 415",	"gr415", 3),
("DECOR TONAL ECREME",	"cRm", 3)
